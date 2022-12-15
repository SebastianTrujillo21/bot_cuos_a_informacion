from encodings import utf_8
import time
import numpy as np
import pandas as pd
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

encoding: utf_8
filename = 'CUPS.xlsx'
filesheet = pd.read_excel(filename, usecols='A')

workbook = load_workbook(filename)
worksheet = workbook.active

cup = list(filesheet.loc[:, 'CUPS_DE_LUZ'])
calle = []
cp = []
localidad = []
provincia = []
cup_de_gas = []


class webScrapping:
    def recolectar_info(self):
        for i in range(2):
            try:
                driver.find_element(By.XPATH, "//body/div/section[@role='main']/div/div/div/div/div/div/div/div/div/div/div/div/div/input[1]").send_keys(cup[i])
                time.sleep(2)
                driver.find_element(By.XPATH,
                                    "//body[1]/div[1]/section[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/button[1]").click()
                time.sleep(40)
                box_direccion = driver.find_element(By.XPATH,
                                                    "//body[1]/div[1]/section[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div["
                                                    "2]/div[1]/p[1]")               #RECOGE TODO LOS VALORES DE LA INFORMACION QUE DAL EL CUP EN BOX
            except NoSuchElementException:
                driver.refresh()
                continue
            time.sleep(5)
            list_direccion = box_direccion.text.split("\n")                 #LOS SEPARA EN UNA LISTA POR CADA ESTA QUE HAY ['CALLE BUEN RETIRO, 11 3 9', '46015 - VALÈNCIA', 'VALENCIA/VALÈNCIA', 'CUPS :ES 0021 **** **** 1068 MX']
            list_direccion.pop()                                            #ELIMINAMOS EL VALOR DE CUP QUE NO NOS SIRVE ['CALLE BUEN RETIRO, 11 3 9', '46015 - VALÈNCIA', 'VALENCIA/VALÈNCIA']
            arr_direccion = np.array(list_direccion)
            calle.append(arr_direccion[0])

            a, b = str(arr_direccion[1]).split("-", 1)                      #DIVIDIMOS LA LOVALIDAD Y EL CODIGO POSTAL
            cp.append(str(a).strip())
            localidad.append(str(b).strip())
            provincia.append(arr_direccion[2])                              #AÑADIMOS LA PROVINCIA

            try:
                driver.find_element(By.XPATH, "//div[@id='gas']").click()
            except NoSuchElementException:
                driver.refresh()
                continue
            time.sleep(20)

            try:
                cup_gas = driver.find_element(By.XPATH,
                                              "//div[@class='MuiBox-root jss81']//p[1]").text
                cup_de_gas.append(cup_gas[5:].strip())                      #Añadimos el cup de gas :D

            except NoSuchElementException:
                driver.refresh()
                continue

            for a, value in enumerate(calle):
                worksheet.cell(row=i + 2, column=2, value=value)
            for a, value in enumerate(cp):
                worksheet.cell(row=i + 2, column=3, value=value)
            for a, value in enumerate(localidad):
                worksheet.cell(row=i + 2, column=4, value=value)
            for a, value in enumerate(provincia):
                worksheet.cell(row=i + 2, column=5, value=value)
            for a, value in enumerate(cup_de_gas):
                worksheet.cell(row=i + 2, column=6, value=value)
            driver.refresh()


url = 'https://front-calculator.zapotek.adn.naturgy.com/'
driver = webdriver.Chrome()
driver.maximize_window()
driver.get(url)
trabajo = webScrapping()
trabajo.recolectar_info()
workbook.save(filename)
